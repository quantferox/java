"""
Excel reference for Spring Boot annotations.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, Rule
from openpyxl.styles.differential import DifferentialStyle

# ─── Data — logical learning order ───────────────────────────────────────────
# (group_title, [ (annotation, in_project), ... ])
# Single group color for all headers — deep navy

GROUP_COLOR = "1B3A6B"   # one color for ALL group headers

GROUPS = [
    ("Java Standard", [
        ("@Override",            True),
        ("@Deprecated",          False),
        ("@SuppressWarnings",    False),
        ("@FunctionalInterface", False),
        ("@SafeVarargs",         False),
    ]),
    ("Spring Core — Stereotypes", [
        ("@SpringBootApplication",    True),
        ("@Component",                True),
        ("@Service",                  True),
        ("@Repository",               True),
        ("@Controller",               True),
        ("@RestController",           True),
        ("@Configuration",            True),
        ("@ControllerAdvice",         True),
        ("@RestControllerAdvice",     True),
    ]),
    ("Spring Core — Bean Lifecycle", [
        ("@Bean",          True),
        ("@Autowired",     True),
        ("@Qualifier",     False),
        ("@Primary",       False),
        ("@Lazy",          False),
        ("@Scope",         False),
        ("@PostConstruct", False),
        ("@PreDestroy",    False),
    ]),
    ("Spring Core — Configuration", [
        ("@ConfigurationProperties",  True),
        ("@PropertySource",           False),
        ("@Profile",                  False),
        ("@Import",                   False),
        ("@Conditional",              False),
        ("@ConditionalOnProperty",    False),
        ("@ConditionalOnMissingBean", False),
        ("@EnableAsync",              True),
        ("@EnableCaching",            True),
        ("@EnableScheduling",         True),
        ("@EnableJpaAuditing",        True),
        ("@EnableWebSecurity",        True),
        ("@EnableMethodSecurity",     True),
    ]),
    ("Spring Web — Mapping", [
        ("@RequestMapping", True),
        ("@GetMapping",     True),
        ("@PostMapping",    True),
        ("@PutMapping",     True),
        ("@PatchMapping",   True),
        ("@DeleteMapping",  True),
    ]),
    ("Spring Web — Params & Body", [
        ("@PathVariable",   True),
        ("@RequestParam",   True),
        ("@RequestBody",    False),
        ("@RequestHeader",  False),
        ("@CookieValue",    False),
        ("@ModelAttribute", False),
        ("@Valid",          True),
        ("@Validated",      True),
    ]),
    ("Spring Web — Response & Errors", [
        ("@ResponseStatus",          True),
        ("@ResponseBody",            False),
        ("@ExceptionHandler",        True),
        ("@AuthenticationPrincipal", True),
        ("@CrossOrigin",             False),
    ]),
    ("Spring Security", [
        ("@EnableWebSecurity",    True),
        ("@EnableMethodSecurity", True),
        ("@PreAuthorize",         True),
        ("@PostAuthorize",        False),
        ("@PreFilter",            False),
        ("@PostFilter",           False),
        ("@Secured",              False),
        ("@RolesAllowed",         False),
        ("@PermitAll",            False),
        ("@DenyAll",              False),
        ("@WithMockUser",         False),
    ]),
    ("Spring Data JPA", [
        ("@Repository",       True),
        ("@Query",            True),
        ("@Param",            True),
        ("@Modifying",        True),
        ("@Transactional",    True),
        ("@Lock",             False),
        ("@EntityGraph",      False),
        ("@NoRepositoryBean", False),
    ]),
    ("JPA / Hibernate — Entity", [
        ("@Entity",           True),
        ("@Table",            True),
        ("@MappedSuperclass", True),
        ("@Embeddable",       False),
        ("@Embedded",         False),
        ("@EmbeddedId",       False),
        ("@EntityListeners",  True),
        ("@NamedQuery",       False),
    ]),
    ("JPA / Hibernate — Columns & ID", [
        ("@Id",                True),
        ("@GeneratedValue",    True),
        ("@SequenceGenerator", False),
        ("@Column",            True),
        ("@Lob",               True),
        ("@Version",           True),
        ("@Enumerated",        True),
        ("@Transient",         False),
        ("@Formula",           False),
        ("@Type",              False),
    ]),
    ("JPA / Hibernate — Constraints", [
        ("@UniqueConstraint", True),
        ("@Index",            True),
        ("@Check",            False),
    ]),
    ("JPA / Hibernate — Relations", [
        ("@OneToOne",   False),
        ("@OneToMany",  True),
        ("@ManyToOne",  True),
        ("@ManyToMany", False),
        ("@JoinColumn", True),
        ("@JoinTable",  False),
        ("@OrderBy",    False),
    ]),
    ("JPA / Hibernate — Lifecycle Callbacks", [
        ("@PrePersist",  True),
        ("@PostPersist", False),
        ("@PreUpdate",   False),
        ("@PostUpdate",  False),
        ("@PreRemove",   False),
        ("@PostRemove",  False),
        ("@PostLoad",    False),
    ]),
    ("JPA / Hibernate — Auditing", [
        ("@CreatedDate",      True),
        ("@LastModifiedDate", True),
        ("@CreatedBy",        False),
        ("@LastModifiedBy",   False),
    ]),
    ("Bean Validation — Nullability", [
        ("@NotNull",  True),
        ("@NotBlank", True),
        ("@NotEmpty", True),
        ("@Null",     True),
    ]),
    ("Bean Validation — Numeric", [
        ("@Min",            True),
        ("@Max",            False),
        ("@DecimalMin",     True),
        ("@DecimalMax",     False),
        ("@Digits",         True),
        ("@Positive",       True),
        ("@PositiveOrZero", False),
        ("@Negative",       False),
        ("@NegativeOrZero", False),
    ]),
    ("Bean Validation — String & Date", [
        ("@Size",            True),
        ("@Email",           True),
        ("@Pattern",         False),
        ("@Past",            False),
        ("@PastOrPresent",   False),
        ("@Future",          False),
        ("@FutureOrPresent", False),
        ("@Constraint",      False),
        ("@GroupSequence",   False),
    ]),
    ("Spring Cache", [
        ("@Cacheable",   True),
        ("@CacheEvict",  True),
        ("@CachePut",    False),
        ("@Caching",     True),
        ("@CacheConfig", False),
    ]),
    ("Spring Async & Scheduling", [
        ("@Async",     True),
        ("@Scheduled", True),
    ]),
    ("Spring Events", [
        ("@EventListener",              True),
        ("@TransactionalEventListener", False),
    ]),
    ("Lombok", [
        ("@Data",                    True),
        ("@Value",                   False),
        ("@Builder",                 True),
        ("@SuperBuilder",            False),
        ("@Getter",                  True),
        ("@Setter",                  True),
        ("@ToString",                False),
        ("@EqualsAndHashCode",       False),
        ("@NoArgsConstructor",       True),
        ("@AllArgsConstructor",      True),
        ("@RequiredArgsConstructor", True),
        ("@Slf4j",                   True),
        ("@Log4j2",                  False),
        ("@NonNull",                 True),
        ("@Accessors",               False),
        ("@With",                    False),
        ("@UtilityClass",            False),
        ("@SneakyThrows",            False),
    ]),
    ("Jackson", [
        ("@JsonInclude",          True),
        ("@JsonProperty",         False),
        ("@JsonAlias",            False),
        ("@JsonIgnore",           False),
        ("@JsonIgnoreProperties", False),
        ("@JsonFormat",           False),
        ("@JsonSerialize",        False),
        ("@JsonDeserialize",      False),
        ("@JsonTypeInfo",         False),
        ("@JsonSubTypes",         False),
        ("@JsonNaming",           False),
        ("@JsonCreator",          False),
        ("@JsonValue",            False),
        ("@JsonManagedReference", False),
        ("@JsonBackReference",    False),
    ]),
    ("MapStruct", [
        ("@Mapper",                      True),
        ("@Mapping",                     True),
        ("@Mappings",                    False),
        ("@MappingTarget",               True),
        ("@BeanMapping",                 True),
        ("@InheritConfiguration",        False),
        ("@InheritInverseConfiguration", False),
        ("@ValueMapping",                False),
        ("@Named",                       False),
    ]),
    ("Swagger / OpenAPI", [
        ("@Tag",                 True),
        ("@Operation",           True),
        ("@Parameter",           False),
        ("@ApiResponse",         False),
        ("@ApiResponses",        False),
        ("@Schema",              False),
        ("@SecurityScheme",      True),
        ("@SecurityRequirement", True),
    ]),
    ("Testing — JUnit 5", [
        ("@Test",              True),
        ("@BeforeEach",        True),
        ("@AfterEach",         False),
        ("@BeforeAll",         False),
        ("@AfterAll",          False),
        ("@DisplayName",       True),
        ("@ExtendWith",        True),
        ("@Nested",            False),
        ("@ParameterizedTest", False),
        ("@ValueSource",       False),
        ("@CsvSource",         False),
        ("@MethodSource",      False),
        ("@EnumSource",        False),
        ("@Timeout",           False),
        ("@Disabled",          False),
        ("@Tag",               False),
        ("@RepeatedTest",      False),
    ]),
    ("Testing — Mockito", [
        ("@Mock",            True),
        ("@InjectMocks",     True),
        ("@Spy",             False),
        ("@Captor",          False),
        ("@MockitoSettings", False),
    ]),
    ("Testing — Spring Test", [
        ("@SpringBootTest",             True),
        ("@WebMvcTest",                 False),
        ("@DataJpaTest",                False),
        ("@DataRedisTest",              False),
        ("@RestClientTest",             False),
        ("@AutoConfigureMockMvc",       True),
        ("@AutoConfigureWebTestClient", False),
        ("@MockBean",                   False),
        ("@SpyBean",                    False),
        ("@ActiveProfiles",             True),
        ("@Sql",                        True),
        ("@DynamicPropertySource",      True),
        ("@Transactional",              True),
        ("@Rollback",                   False),
        ("@TestPropertySource",         False),
    ]),
    ("Testing — Testcontainers", [
        ("@Testcontainers", True),
        ("@Container",      True),
    ]),
    ("Spring Retry", [
        ("@EnableRetry",    False),
        ("@Retryable",      False),
        ("@Recover",        False),
        ("@CircuitBreaker", False),
    ]),
    ("Feign / HTTP Clients", [
        ("@FeignClient",        False),
        ("@EnableFeignClients", False),
    ]),
    ("Spring Actuator / Metrics", [
        ("@Endpoint",        False),
        ("@ReadOperation",   False),
        ("@WriteOperation",  False),
        ("@DeleteOperation", False),
        ("@Timed",           False),
        ("@Counted",         False),
    ]),
    ("Jakarta Standard", [
        ("@PostConstruct", False),
        ("@PreDestroy",    False),
        ("@Resource",      False),
    ]),
]

# ─── Status options ──────────────────────────────────────────────────────────
STATUS_DONE     = "✓ Done"
STATUS_IN_PROG  = "~ In Progress"
STATUS_NOT_DONE = "✗ Not Studied"

DV_FORMULA = f'"{STATUS_DONE},{STATUS_IN_PROG},{STATUS_NOT_DONE}"'

# ─── Style helpers ────────────────────────────────────────────────────────────

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border():
    s = Side(style="thin", color="2C3E50")
    return Border(left=s, right=s, top=s, bottom=s)

ALIGN_C = Alignment(horizontal="center", vertical="center")
ALIGN_L = Alignment(horizontal="left",   vertical="center", indent=1)

BG_DARK    = "0D1117"
BG_ALT     = "161B22"
BG_TITLE   = "0F3460"
BG_HEADER  = "0D2137"

# ─── Builder ─────────────────────────────────────────────────────────────────

def build_excel(out_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Annotations"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "0F3460"
    ws.sheet_view.zoomScale = 115
    ws.freeze_panes = "A3"

    # Column widths: A=# | B=Annotation | C=Status
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 18

    # ── Row 1: main title ────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 40
    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value     = "Spring Boot  —  Annotations Reference"
    t.fill      = fill(BG_TITLE)
    t.font      = Font(name="Segoe UI", bold=True, size=15, color="E8F4FD")
    t.alignment = ALIGN_C

    # ── Row 2: column headers ────────────────────────────────────────────────
    ws.row_dimensions[2].height = 30
    for col, label in enumerate(["#", "Annotation", "Status"], start=1):
        c = ws.cell(row=2, column=col, value=label)
        c.fill      = fill(BG_HEADER)
        c.font      = Font(name="Segoe UI", bold=True, size=11, color="AED6F1")
        c.alignment = ALIGN_C
        c.border    = border()

    # ── Data validation for column C ─────────────────────────────────────────
    dv = DataValidation(
        type="list",
        formula1=DV_FORMULA,
        allow_blank=False,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid",
        error="Pick a value from the list"
    )
    ws.add_data_validation(dv)

    # ── Conditional formatting: color text by status value ───────────────────
    cf_range = "C3:C2000"

    # Done → green
    ws.conditional_formatting.add(cf_range, Rule(
        type="containsText",
        operator="containsText",
        text="Done",
        dxf=DifferentialStyle(font=Font(bold=True, color="27AE60")),
        formula=['NOT(ISERROR(SEARCH("Done",C3)))']
    ))

    # In Progress → blue
    ws.conditional_formatting.add(cf_range, Rule(
        type="containsText",
        operator="containsText",
        text="In Progress",
        dxf=DifferentialStyle(font=Font(bold=True, color="2E86C1")),
        formula=['NOT(ISERROR(SEARCH("In Progress",C3)))']
    ))

    # Not Studied → red
    ws.conditional_formatting.add(cf_range, Rule(
        type="containsText",
        operator="containsText",
        text="Not Studied",
        dxf=DifferentialStyle(font=Font(bold=True, color="E74C3C")),
        formula=['NOT(ISERROR(SEARCH("Not Studied",C3)))']
    ))

    row_num    = 3
    annot_num  = 0
    alt        = False

    for group_title, annotations in GROUPS:

        # ── Group header row (full-width, single color) ───────────────────
        ws.row_dimensions[row_num].height = 26
        ws.merge_cells(f"A{row_num}:C{row_num}")

        gh = ws.cell(row=row_num, column=1, value=f"   {group_title}")
        gh.fill      = fill(GROUP_COLOR)
        gh.font      = Font(name="Segoe UI", bold=True, size=11, color="FFFFFF")
        gh.alignment = Alignment(horizontal="center", vertical="center")
        gh.border    = border()

        for col in (2, 3):
            ws.cell(row=row_num, column=col).fill   = fill(GROUP_COLOR)
            ws.cell(row=row_num, column=col).border = border()

        row_num += 1
        alt = False

        # ── Annotation rows ───────────────────────────────────────────────
        for annotation, _ in annotations:
            annot_num += 1
            bg  = BG_ALT if alt else BG_DARK
            alt = not alt

            ws.row_dimensions[row_num].height = 22

            # A — number
            n = ws.cell(row=row_num, column=1, value=annot_num)
            n.fill      = fill(bg)
            n.font      = Font(name="Segoe UI", size=9, color="4A6FA5")
            n.alignment = ALIGN_C
            n.border    = border()

            # B — annotation name
            a = ws.cell(row=row_num, column=2, value=annotation)
            a.fill      = fill(bg)
            a.font      = Font(name="Consolas", size=10, color="64B5F6")
            a.alignment = ALIGN_L
            a.border    = border()

            # C — status (dropdown)
            status = STATUS_NOT_DONE
            c = ws.cell(row=row_num, column=3, value=status)
            c.fill      = fill(bg)
            c.font      = Font(name="Segoe UI", bold=True, size=10, color="E74C3C")
            c.alignment = ALIGN_C
            c.border    = border()
            dv.add(c)

            row_num += 1

    # ── Footer ───────────────────────────────────────────────────────────────
    ws.row_dimensions[row_num].height = 24
    ws.merge_cells(f"A{row_num}:C{row_num}")
    f_cell = ws.cell(row=row_num, column=1,
                     value=f"Total: {annot_num} annotations")
    f_cell.fill      = fill(BG_TITLE)
    f_cell.font      = Font(name="Segoe UI", bold=True, size=10, color="AED6F1")
    f_cell.alignment = ALIGN_C
    f_cell.border    = border()
    for col in (2, 3):
        ws.cell(row=row_num, column=col).fill   = fill(BG_TITLE)
        ws.cell(row=row_num, column=col).border = border()

    wb.save(out_path)
    print(f"Saved: {out_path}  ({annot_num} annotations, {row_num} rows)")


if __name__ == "__main__":
    build_excel(r"c:\Users\QuantFerox\Desktop\lumeo\annotations-reference-v3.xlsx")
